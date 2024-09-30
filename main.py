import cv2
import numpy as np
import face_recognition
import pickle
from datetime import datetime, date
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.image import Image
from kivy.clock import Clock
from kivy.graphics.texture import Texture
from PIL import Image as PILImage, ImageDraw, ImageFont
import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook, load_workbook
import os


# إعداد الخط العربي
font_path = "C:/Windows/Fonts/arial.ttf"
font = ImageFont.truetype(font_path, 32)

# تحميل ملف التشفيرات
print("Loading Encode File ...")
with open('EncodeFile.pkl', 'rb') as file:
    encodeListKnownWithIds = pickle.load(file)
encodeListKnown, studentIds = encodeListKnownWithIds
print("Encode File Loaded")

class FaceRecognitionApp(App):
    def build(self):
        self.layout = BoxLayout(orientation='vertical')
        
        self.department_input = TextInput(hint_text='enter the section', multiline=False)
        self.layout.add_widget(self.department_input)
        
        self.start_button = Button(text='start', on_press=self.start_recognition)
        self.layout.add_widget(self.start_button)
        
        self.stop_button = Button(text='stop', on_press=self.stop_recognition, disabled=True)
        self.layout.add_widget(self.stop_button)
        
        self.image_widget = Image()
        self.layout.add_widget(self.image_widget)

        self.message_label = Label(text='')
        self.layout.add_widget(self.message_label)
        
        self.capture = None
        self.already_attendence_taken = ""

        return self.layout

    def start_recognition(self, instance):
        self.department = self.department_input.text
        if not self.department:
            self.message_label.text = "please enter the section of solider"
            return

        self.start_button.disabled = True
        self.stop_button.disabled = False

        # إعداد الكاميرا
        self.capture = cv2.VideoCapture(0)
        self.capture.set(3, 680)
        self.capture.set(4, 780)
        
        # بدء التعرف على الوجوه
        Clock.schedule_interval(self.update, 1.0 / 30.0)

    def update(self, dt):
        success, img = self.capture.read()
        if not success:
            return

        imgS = cv2.resize(img, (0 , 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        faceCurFrame = face_recognition.face_locations(imgS)
        encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

        for encodeFace, faceLoc in zip(encodeCurFrame, faceCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace, tolerance=0.4)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            best_match_index = np.argmin(faceDis)
            id = 'Unknown'
            if matches[best_match_index] and faceDis[best_match_index] < 0.4:
                strict_match = face_recognition.compare_faces([encodeListKnown[best_match_index]], encodeFace, tolerance=0.35)
                if strict_match[0]:
                    id = studentIds[best_match_index]

            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)

            img = self.draw_text_arabic(img, str(id), (x1, y2 + 20), font, (0, 255, 0))

            if self.already_attendence_taken != id and id != "Unknown":
                current_time = datetime.now().strftime('%H:%M:%S')
                self.save_to_excel(id, current_time, self.department)
                
                # إعادة تشكيل النص العربي
                reshaped_id = arabic_reshaper.reshape(id)
                bidi_id = get_display(reshaped_id)
                self.message_label.text = f"ok recognation: {bidi_id} in {current_time}"
                self.already_attendence_taken = id

        buf = cv2.flip(img, 0).tostring()
        texture = Texture.create(size=(img.shape[1], img.shape[0]), colorfmt='bgr')
        texture.blit_buffer(buf, colorfmt='bgr', bufferfmt='ubyte')
        self.image_widget.texture = texture

    def draw_text_arabic(self, image, text, position, font, color=(0, 255, 0)):
        reshaped_text = arabic_reshaper.reshape(text)
        bidi_text = get_display(reshaped_text)
        img_pil = PILImage.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
        draw = ImageDraw.Draw(img_pil)
        draw.text(position, bidi_text, font=font, fill=color)
        return cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)

    def save_to_excel(self, id, time, department):
        excel_file = 'attendance.xlsx'
    
    # التحقق مما إذا كان ملف Excel موجودًا بالفعل
        if not os.path.exists(excel_file):
        # إنشاء ملف Excel جديد إذا لم يكن موجودًا
           self.workbook = Workbook()
           self.sheet = self.workbook.active
           self.sheet.title = "Attendance"
           self.sheet.append(["ID", "Time", "Section","Date"])
        else:
        # فتح ملف Excel إذا كان موجودًا
           self.workbook = load_workbook(excel_file)
           self.sheet = self.workbook.active
    
    # إضافة سجل جديد إلى الجدول
        self.sheet.append([id, time, department])
    
    # حفظ التعديلات
        self.workbook.save(excel_file)


    def stop_recognition(self, instance):
        self.start_button.disabled = False
        self.stop_button.disabled = True
        if self.capture:
            self.capture.release()
        Clock.unschedule(self.update)
        self.message_label.text ="stoped succssfully"

if __name__ == '__main__':
    FaceRecognitionApp().run()

